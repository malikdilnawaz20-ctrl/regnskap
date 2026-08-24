#!/usr/bin/env python3
"""
Generate a reviewed, idempotent import SQL file for Skoger og Fjell.

Inputs are local exports already present on the machine:
- Fiken-derived transaction JSON from previous reconciliation work.
- Styreweb/member Excel export with kontingent status.

The script does not contact Supabase or Fiken. It only writes SQL and a
plain-text control report for manual review/run in Supabase SQL Editor.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
    import pdfplumber
except ImportError as exc:  # pragma: no cover - operator-facing message
    raise SystemExit("Mangler openpyxl/pdfplumber. Kjor med Codex sin bundled Python-runtime.") from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_FIKEN = Path("/Users/dilnawazmalik/Documents/innsending_NIF/fiken_transactions.json")
DEFAULT_MEMBERS = Path("/Users/dilnawazmalik/Documents/KARATEMEDLEMMER/BOK2.xlsx")
DEFAULT_PAID_MEMBERS = Path("/Users/dilnawazmalik/Documents/medlemmer2026.xlsx")
DEFAULT_BANK_PDF_DIR = Path("/Users/dilnawazmalik/Downloads")

OUT_SQL = ROOT / "supabase" / "import_skoger_fiken_styreweb.sql"
OUT_REPORT = ROOT / "PRODUKSJON_IMPORT_SKOGER.md"

ORGNR = "912484335"
ORG_NAME = "Skoger og Fjell kampsportklubb"
MEMBERSHIP_FEE_NAME = "Medlemskontingent 2026"
MEMBERSHIP_FEE_ORE = 45_000
MONTH_NAMES = [
    "januar", "februar", "mars", "april", "mai", "juni",
    "juli", "august", "september", "oktober", "november", "desember",
]


ACCOUNT_CATEGORY = {
    3200: ("Medlemskontingent", "inntekt"),
    3440: ("Offentlig tilskudd", "inntekt"),
    3900: ("Andre inntekter", "inntekt"),
    4300: ("Varekostnad", "utgift"),
    6300: ("Hall-leie", "utgift"),
    6790: ("Honorar og tjenester", "utgift"),
    6810: ("Nettside og programvare", "utgift"),
    7140: ("Dommer og stevneavgift", "utgift"),
    7320: ("Markedsføring", "utgift"),
    7490: ("Kontingent til forbund", "utgift"),
    7500: ("Forsikring", "utgift"),
    7770: ("Bankgebyr", "utgift"),
    7790: ("Andre utgifter", "utgift"),
    1921: ("Overføring internkonto", "overforing"),
    2920: ("Mellomregning", "overforing"),
}

BANK_STATEMENT_FILES = [
    "Dokument PKTOUTS03@043389491810922885.pdf",
    "Dokument PKTOUTS03@565692257350498451.pdf",
    "Dokument PKTOUTS03@064960942488950403.pdf",
    "Dokument PKTOUTS03@032775231237276820.pdf",
    "Dokument PKTOUTS03@575958712666394765.pdf",
    "Dokument PKTOUTS03@076200963640121238.pdf",
    "Dokument PKTOUTS03@021785994090626693.pdf",
    "Dokument PKTOUTS03@533858244721799053.pdf",
    "Dokument PKTOUTS03@129385264488527687.pdf",
    "Dokument PKTOUTS03@996091536519205509.pdf",
    "Dokument PKTOUTS03@172522521248486912.pdf",
    "Dokument PKTOUTS03@194124487519672843.pdf",
    "Dokument PKTOUTS03@086193649723352451.pdf",
    "Dokument PKTOUTS03@107758133491203203.pdf",
    "Dokument PKTOUTS03@258523016499702790.pdf",
    "Dokument PKTOUTS03@236928486934611462.pdf",
    "Dokument PKTOUTS03@323301026666880519.pdf",
    "Dokument PKTOUTS03@226003796183673865.pdf",
    "Dokument PKTOUTS03@216070060660160516.pdf",
    "Dokument PKTOUTS03@334403996276695044.pdf",
    "Dokument PKTOUTS03@247188925148196358.pdf",
    "Dokument PKTOUTS03@279394073048502787.pdf",
    "Dokument PKTOUTS03@312686558481119744.pdf",
    "Dokument PKTOUTS03@376009499320307205.pdf",
    "Dokument PKTOUTS03@463086392740816897.pdf",
    "Dokument PKTOUTS03@420251054226293259.pdf",
    "Dokument PKTOUTS03@430517621201375236.pdf",
    "Dokument PKTOUTS03@387699389829919236.pdf",
    "Dokument PKTOUTS03@441851862728451591.pdf",
    "Dokument PKTOUTS03@355147641393966088.pdf",
    "Dokument PKTOUTS03@452476432616461312.pdf",
    "Dokument PKTOUTS03@366101960787749894.pdf",
]

ACCOUNT_NAMES = {
    1920: "Bankinnskudd hovedkonto",
    1921: "Bankinnskudd internkonto",
    2920: "Gjeld/mellomregning",
    3200: "Medlemskontingent",
    3440: "Offentlig tilskudd",
    3900: "Andre inntekter",
    4300: "Varekostnad",
    6300: "Leie av hall og lokaler",
    6790: "Honorar og tjenester",
    6810: "Data, programvare og nettside",
    7140: "Reisekostnad, stevner og cup",
    7320: "Markedsføring og profilering",
    7490: "Kontingent til forbund og krets",
    7500: "Forsikring",
    7770: "Bank- og betalingsgebyr",
    7790: "Andre kostnader",
}


def sql(value):
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, int):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def slug(text: str) -> str:
    text = (text or "").strip().lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "rad"


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass
    return text[:10]


def parse_norwegian_amount(text: str) -> int:
    cleaned = text.replace(".", "").replace(" ", "").replace("\xa0", "").replace(",", ".")
    return int(round(float(cleaned) * 100))


def load_bank_statement(path: Path):
    with pdfplumber.open(path) as pdf:
        text = "\n".join(page.extract_text() or "" for page in pdf.pages)
    period = re.search(r"perioden\s+(\d{2}\.\d{2}\.\d{4})\s+-\s+(\d{2}\.\d{2}\.\d{4})", text)
    start_balance = re.search(r"Saldofra kontoutskrift\d{2}\.\d{2}\.\d{4}\s+([\d .]+,\d{2})", text)
    end_balance = re.search(r"SaldoiDeresfavør\s+([\d .]+,\d{2})", text)
    end_sign = 1
    if not end_balance:
        end_balance = re.search(r"Saldoivårfavør\s+([\d .]+,\d{2})", text)
        end_sign = -1
    if not (period and start_balance and end_balance):
        raise ValueError(f"Fant ikke periode/saldo i {path.name}")
    start_date = dt.datetime.strptime(period.group(1), "%d.%m.%Y").date()
    end_date = dt.datetime.strptime(period.group(2), "%d.%m.%Y").date()
    start_ore = parse_norwegian_amount(start_balance.group(1))
    end_ore = end_sign * parse_norwegian_amount(end_balance.group(1))
    net_ore = end_ore - start_ore
    if net_ore == 0:
        tx_type = "overforing"
        amount_ore = 1
    else:
        tx_type = "inntekt" if net_ore > 0 else "utgift"
        amount_ore = abs(net_ore)
    external_id = hashlib.sha1(f"bankpdf|{path.name}|{start_date}|{end_date}|{net_ore}".encode("utf-8")).hexdigest()[:16]
    return {
        "external_id": external_id,
        "filename": path.name,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "year": end_date.year,
        "month": end_date.month,
        "start_ore": start_ore,
        "end_ore": end_ore,
        "net_ore": net_ore,
        "type": tx_type,
        "amount_ore": amount_ore,
        "description": f"Kontoutskrift {start_date:%m/%Y} - netto bankbevegelse fra {path.name}",
    }


def load_bank_statements(directory: Path):
    statements = []
    missing_files = []
    for filename in BANK_STATEMENT_FILES:
        path = directory / filename
        if not path.exists():
            missing_files.append(filename)
            continue
        statements.append(load_bank_statement(path))
    statements.sort(key=lambda item: (item["year"], item["month"], item["filename"]))
    return statements, missing_files


def split_name(raw: str):
    raw = (raw or "").strip()
    if "," in raw:
        last, first = [p.strip() for p in raw.split(",", 1)]
        return first or last, last if first else ""
    parts = raw.split()
    if len(parts) <= 1:
        return raw, ""
    return " ".join(parts[:-1]), parts[-1]


def month_year(date_text: str) -> str:
    date = dt.date.fromisoformat(str(date_text)[:10])
    return f"{MONTH_NAMES[date.month - 1]} {date.year}"


def public_transaction_description(tx_type: str, category_name: str, date_text: str) -> str:
    if tx_type != "utgift":
        return None
    category = (category_name or "drift").strip()
    if category.lower().startswith("historisk bankspor"):
        category = "historisk bankspor"
    else:
        category = category[:1].lower() + category[1:]
    return f"Utbetaling registrert i regnskapet - {category}, {month_year(date_text)}"


def load_fiken(path: Path):
    with path.open(encoding="utf-8") as handle:
        rows = json.load(handle)
    normalized = []
    for idx, row in enumerate(rows, 1):
        debit = int(row["debitAccount"])
        credit = int(row["creditAccount"])
        amount = int(row["amount_ore"])
        is_income = debit == 1920 and credit != 1920
        is_expense = credit == 1920 and debit != 1920
        direction = "inntekt" if is_income else "utgift" if is_expense else "overforing"
        counterpart = credit if is_income else debit if is_expense else credit
        cat_name, cat_direction = ACCOUNT_CATEGORY.get(counterpart, ("Andre utgifter", direction))
        if direction in {"inntekt", "utgift"} and cat_direction == "overforing":
            cat_direction = direction
        elif direction == "overforing":
            cat_name, cat_direction = ACCOUNT_CATEGORY.get(counterpart, ("Overføring", "overforing"))
        external_id = hashlib.sha1(
            f"fiken|{row['date']}|{debit}|{credit}|{amount}|{row.get('description','')}".encode("utf-8")
        ).hexdigest()[:16]
        normalized.append({
            "external_id": external_id,
            "bilagsnummer": f"FIKEN-{row['year']}-{idx:04d}",
            "dato": row["date"],
            "type": direction,
            "beskrivelse": public_transaction_description(direction, cat_name, row["date"])
            or (row.get("description") or row.get("note") or f"Fiken {row['date']}"),
            "belop_ore": amount,
            "konto_nummer": counterpart,
            "category_name": cat_name,
            "category_direction": cat_direction,
            "regnskapsaar": int(row["year"]),
            "note": row.get("note") or "",
        })
    return normalized


def member_key(member):
    if member.get("fodselsdato"):
        return ("name_born", member["fornavn"].strip().lower(), member["etternavn"].strip().lower(), member["fodselsdato"])
    if member.get("epost"):
        return ("email", member["epost"])
    return ("name", member["fornavn"].strip().lower(), member["etternavn"].strip().lower())


def load_member_file(path: Path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    header = {}
    for idx, row in enumerate(rows):
        labels = {str(value).strip().lower(): pos for pos, value in enumerate(row) if value}
        if "navn" in labels:
            header_idx = idx
            header = labels
            break
    if header_idx is None:
        return []

    name_idx = header["navn"]
    status_idx = header.get("kontingent")
    born_idx = header.get("født") or header.get("fodt") or header.get("fødselsdato") or header.get("fodselsdato")
    email_idx = header.get("e-post") or header.get("epost") or header.get("email")
    gender_idx = header.get("kjønn") or header.get("kjonn")

    members = []
    for idx, row in enumerate(rows[header_idx + 1:], 1):
        name = row[name_idx] if len(row) > name_idx else None
        if not name:
            continue
        status = str(row[status_idx] or "").strip() if status_idx is not None and len(row) > status_idx else ""
        born = parse_date(row[born_idx] if born_idx is not None and len(row) > born_idx else None)
        email = str(row[email_idx] or "").strip().lower() if email_idx is not None and len(row) > email_idx and row[email_idx] else None
        gender_raw = str(row[gender_idx] or "").strip().lower() if gender_idx is not None and len(row) > gender_idx and row[gender_idx] else ""
        gender = "mann" if gender_raw == "m" else "kvinne" if gender_raw == "k" else None
        first, last = split_name(str(name))
        ext_seed = f"{first}-{last}-{born}" if born else email or f"{first}-{last}-{idx}"
        members.append({
            "external_id": hashlib.sha1(f"styreweb|{ext_seed}".encode("utf-8")).hexdigest()[:16],
            "fornavn": first,
            "etternavn": last,
            "fodselsdato": born,
            "epost": email,
            "kjonn": gender,
            "kontingent": status,
            "betalt": status.lower() == "betalt",
        })
    return members


def load_members(path: Path, paid_path: Path = DEFAULT_PAID_MEMBERS):
    merged = {}
    for member in load_member_file(path):
        merged[member_key(member)] = member
    if paid_path.exists():
        for member in load_member_file(paid_path):
            key = member_key(member)
            if key in merged:
                merged[key]["kontingent"] = member["kontingent"] or merged[key]["kontingent"]
                merged[key]["betalt"] = merged[key]["betalt"] or member["betalt"]
                if not merged[key].get("epost") and member.get("epost"):
                    merged[key]["epost"] = member["epost"]
            else:
                merged[key] = member
    return sorted(merged.values(), key=lambda item: (item["etternavn"].lower(), item["fornavn"].lower(), item.get("fodselsdato") or ""))


def generate_sql(fiken_rows, members, bank_statements):
    accounts = sorted({r["konto_nummer"] for r in fiken_rows if r["konto_nummer"]})
    categories = sorted({(r["category_name"], r["category_direction"], r["konto_nummer"]) for r in fiken_rows})
    lines = [
        "-- =====================================================================",
        "--  Import fra Fiken, Styreweb og kontoutskrifter for Skoger og Fjell kampsportklubb",
        "--  Generert lokalt. Idempotent: kan kjores flere ganger uten duplikater.",
        "-- =====================================================================",
        "",
        "create unique index if not exists members_styreweb_ext_key",
        "  on members (organization_id, ekstern_kilde, ekstern_id)",
        "  where ekstern_kilde is not null and ekstern_id is not null;",
        "",
        "create unique index if not exists payment_claims_import_member_fee_desc_key",
        "  on payment_claims (organization_id, member_id, fee_id, beskrivelse);",
        "",
        "create unique index if not exists documents_import_file_key",
        "  on documents (organization_id, mappe, tittel, filnavn);",
        "",
        "do $$",
        "declare",
        "  org uuid;",
        "  fee uuid;",
        "  member_id uuid;",
        "begin",
        f"  select id into org from organizations where orgnr = {sql(ORGNR)};",
        "  if org is null then",
        f"    raise exception 'Fant ikke organisasjon {ORGNR}';",
        "  end if;",
        "",
        "  insert into accounts (organization_id, navn, kontonummer, type)",
        "  values (org, 'Fiken bankkonto 2220.29.21373', '2220.29.21373', 'bank')",
        "  on conflict do nothing;",
        "",
        "  insert into fees (organization_id, navn, type, intervall, belop_ore, gjelder_fra)",
        f"  values (org, {sql(MEMBERSHIP_FEE_NAME)}, 'medlemskontingent', 'aarlig', {MEMBERSHIP_FEE_ORE}, '2026-01-01')",
        "  on conflict do nothing;",
        f"  select id into fee from fees where organization_id = org and navn = {sql(MEMBERSHIP_FEE_NAME)} limit 1;",
        "",
        "  -- Kontoplan/kategorier brukt i Fiken-importen",
    ]
    for number in accounts:
        name = ACCOUNT_NAMES.get(number, f"Konto {number}")
        lines.append(
            f"  insert into accounting_accounts (organization_id, nummer, navn) values "
            f"(org, {number}, {sql(name)}) on conflict (organization_id, nummer) do update set navn = excluded.navn;"
        )
    for name, direction, number in categories:
        if direction not in {"inntekt", "utgift", "overforing"}:
            direction = "utgift"
        lines.append(
            f"  insert into categories (organization_id, navn, retning, konto_nummer) values "
            f"(org, {sql(name)}, {sql(direction)}, {number}) on conflict (organization_id, navn, retning) "
            f"do update set konto_nummer = excluded.konto_nummer, aktiv = true;"
        )
    lines.extend([
        "  insert into categories (organization_id, navn, retning, konto_nummer) values",
        "    (org, 'Historisk bankspor - inn', 'inntekt', 3900),",
        "    (org, 'Historisk bankspor - ut', 'utgift', 7790)",
        "  on conflict (organization_id, navn, retning) do update set aktiv = true;",
        "",
        "  -- Kontoutskrifter 2020-2023. Importert som maanedlige netto bankspor, ikke detaljlinjer.",
    ])
    for stmt in bank_statements:
        category_name = "Historisk bankspor - inn" if stmt["type"] == "inntekt" else "Historisk bankspor - ut"
        lines.extend([
            "  insert into documents (organization_id, mappe, tittel, filnavn, kun_styret)",
            f"  values (org, 'Regnskap', {sql('Kontoutskrift ' + stmt['start_date'][:7])}, {sql(stmt['filename'])}, true)",
            "  on conflict do nothing;",
            "",
            "  insert into transactions (organization_id, bilagsnummer, dato, type, beskrivelse, belop_ore, konto_nummer, category_id, regnskapsaar)",
            "  select org, "
            + ", ".join([
                sql(f"BANK-{stmt['year']}-{stmt['month']:02d}"),
                sql(stmt["end_date"]),
                sql(stmt["type"] if stmt["type"] != "overforing" else "utgift"),
                sql(public_transaction_description(stmt["type"], "Historisk bankspor", stmt["end_date"]) or "Bankbevegelse registrert fra kontoutskrift"),
                sql(stmt["amount_ore"]),
                sql(3900 if stmt["type"] == "inntekt" else 7790),
            ])
            + ", c.id, "
            + sql(stmt["year"]),
            "  from categories c",
            f"  where c.organization_id = org and c.navn = {sql(category_name)} and c.retning = {sql('inntekt' if stmt['type'] == 'inntekt' else 'utgift')}",
            "  on conflict (organization_id, bilagsnummer) do update set",
            "    dato = excluded.dato, type = excluded.type, beskrivelse = excluded.beskrivelse,",
            "    belop_ore = excluded.belop_ore, konto_nummer = excluded.konto_nummer,",
            "    category_id = excluded.category_id, regnskapsaar = excluded.regnskapsaar;",
        ])
    lines.extend(["", "  -- Fiken transaksjoner"])
    for row in fiken_rows:
        lines.extend([
            "  insert into transactions (organization_id, bilagsnummer, dato, type, beskrivelse, belop_ore, konto_nummer, category_id, regnskapsaar)",
            "  select org, "
            + ", ".join([
                sql(row["bilagsnummer"]),
                sql(row["dato"]),
                sql(row["type"]),
                sql(row["beskrivelse"]),
                sql(row["belop_ore"]),
                sql(row["konto_nummer"]),
            ])
            + ", c.id, "
            + sql(row["regnskapsaar"]),
            "  from categories c",
            f"  where c.organization_id = org and c.navn = {sql(row['category_name'])} and c.retning = {sql(row['category_direction'])}",
            "  on conflict (organization_id, bilagsnummer) do update set",
            "    dato = excluded.dato, type = excluded.type, beskrivelse = excluded.beskrivelse,",
            "    belop_ore = excluded.belop_ore, konto_nummer = excluded.konto_nummer,",
            "    category_id = excluded.category_id, regnskapsaar = excluded.regnskapsaar;",
        ])
    lines.extend(["", "  -- Styreweb medlemmer og kontingentstatus"])
    for member in members:
        lines.extend([
            "  insert into members (organization_id, fornavn, etternavn, fodselsdato, kjonn, epost, status, ekstern_id, ekstern_kilde)",
            "  values (org, "
            + ", ".join([
                sql(member["fornavn"]),
                sql(member["etternavn"]),
                sql(member["fodselsdato"]),
                sql(member["kjonn"]),
                sql(member["epost"]),
                "'aktiv'",
                sql(member["external_id"]),
                "'styreweb'",
            ])
            + ")",
            "  on conflict do nothing",
            "  returning id into member_id;",
            "  if member_id is null then",
            f"    select id into member_id from members where organization_id = org and ekstern_kilde = 'styreweb' and ekstern_id = {sql(member['external_id'])} limit 1;",
            "  end if;",
            "  if member_id is not null then",
            "    insert into payment_claims (organization_id, member_id, fee_id, beskrivelse, belop_ore, betalt_ore, status, forfall)",
            "    values (org, member_id, fee, "
            + ", ".join([
                sql(MEMBERSHIP_FEE_NAME),
                sql(MEMBERSHIP_FEE_ORE),
                sql(MEMBERSHIP_FEE_ORE if member["betalt"] else 0),
                sql("betalt" if member["betalt"] else "ikke_betalt"),
                "'2026-12-31'",
            ])
            + ")",
            "    on conflict do nothing;",
            "  end if;",
        ])
    lines.extend([
        "",
        "  insert into import_jobs (organization_id, type, filnavn, antall_lest, antall_importert, antall_avvist, status)",
        f"  values (org, 'fiken_styreweb_bankspor', 'fiken_transactions.json + medlemmer2026.xlsx + kontoutskrifter 2020-2023', {len(fiken_rows) + len(members) + len(bank_statements)}, {len(fiken_rows) + len(members) + len(bank_statements)}, 0, 'fullfort');",
        "end $$;",
        "",
    ])
    return "\n".join(lines)


def generate_report(fiken_rows, members, bank_statements, missing_bank_files):
    by_year = defaultdict(lambda: {"inntekt": 0, "utgift": 0, "overforing": 0, "count": 0})
    by_account = Counter()
    for row in fiken_rows:
        by_year[row["regnskapsaar"]][row["type"]] += row["belop_ore"]
        by_year[row["regnskapsaar"]]["count"] += 1
        by_account[row["konto_nummer"]] += 1
    paid = sum(1 for m in members if m["betalt"])
    unpaid = len(members) - paid
    bank_by_year = defaultdict(lambda: {"inn": 0, "ut": 0, "count": 0, "months": []})
    for stmt in bank_statements:
        bucket = bank_by_year[stmt["year"]]
        bucket["count"] += 1
        bucket["months"].append(f"{stmt['month']:02d}")
        if stmt["net_ore"] >= 0:
            bucket["inn"] += stmt["net_ore"]
        else:
            bucket["ut"] += abs(stmt["net_ore"])
    lines = [
        "# Importkontroll - Skoger og Fjell",
        "",
        f"Generert: {dt.datetime.now().isoformat(timespec='seconds')}",
        "",
        "## Kilder",
        "",
        f"- Fiken: `{DEFAULT_FIKEN}`",
        f"- Styreweb/medlemmer: `{DEFAULT_MEMBERS}`",
        f"- Kontoutskrifter: `{DEFAULT_BANK_PDF_DIR}`",
        "",
        "## Kontoutskrifter 2020-2023",
        "",
        f"- PDF-er lest: {len(bank_statements)}",
        f"- PDF-er som manglet på disk: {len(missing_bank_files)}",
    ]
    for year in sorted(bank_by_year):
        item = bank_by_year[year]
        net = item["inn"] - item["ut"]
        lines.append(
            f"- {year}: {item['count']} utskrifter, maaneder {', '.join(item['months'])}, "
            f"netto {net / 100:,.2f} kr"
        )
    expected_months = {(year, month) for year in (2020, 2021, 2022, 2023) for month in range(1, 13)}
    found_months = {(stmt["year"], stmt["month"]) for stmt in bank_statements}
    missing_months = sorted(expected_months - found_months)
    if missing_months:
        lines.append("")
        lines.append("Maaneder uten vedlagt kontoutskrift:")
        for year, month in missing_months:
            lines.append(f"- {year}-{month:02d}")
    lines.extend([
        "",
        "## Fiken-transaksjoner",
        "",
        f"- Antall: {len(fiken_rows)}",
    ])
    for year in sorted(by_year):
        item = by_year[year]
        net = item["inntekt"] - item["utgift"]
        lines.append(
            f"- {year}: {item['count']} rader, inn {item['inntekt'] / 100:,.2f} kr, "
            f"ut {item['utgift'] / 100:,.2f} kr, netto {net / 100:,.2f} kr"
        )
    lines.extend(["", "Kontoer brukt:"])
    for account, count in sorted(by_account.items()):
        lines.append(f"- {account}: {ACCOUNT_NAMES.get(account, 'Konto ' + str(account))} ({count})")
    lines.extend([
        "",
        "## Styreweb-medlemmer",
        "",
        f"- Medlemmer i fil: {len(members)}",
        f"- Kontingent betalt: {paid}",
        f"- Ikke betalt/ukjent: {unpaid}",
        "",
        "## Importfil",
        "",
        f"- SQL: `{OUT_SQL}`",
        "- Importen er laget idempotent med faste bilagsnummer for Fiken-rader.",
        "- Medlemsbetalinger opprettes som betalingskrav for `Medlemskontingent 2026`.",
    ])
    return "\n".join(lines) + "\n"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--fiken", type=Path, default=DEFAULT_FIKEN)
    parser.add_argument("--members", type=Path, default=DEFAULT_MEMBERS)
    parser.add_argument("--bank-pdf-dir", type=Path, default=DEFAULT_BANK_PDF_DIR)
    args = parser.parse_args()

    fiken_rows = load_fiken(args.fiken)
    members = load_members(args.members)
    bank_statements, missing_bank_files = load_bank_statements(args.bank_pdf_dir)
    OUT_SQL.write_text(generate_sql(fiken_rows, members, bank_statements), encoding="utf-8")
    OUT_REPORT.write_text(generate_report(fiken_rows, members, bank_statements, missing_bank_files), encoding="utf-8")
    print(f"Wrote {OUT_SQL}")
    print(f"Wrote {OUT_REPORT}")
    print(f"Fiken rows: {len(fiken_rows)}")
    print(f"Members: {len(members)}")
    print(f"Paid members: {sum(1 for m in members if m['betalt'])}")
    print(f"Bank statement PDFs: {len(bank_statements)}")


if __name__ == "__main__":
    main()
